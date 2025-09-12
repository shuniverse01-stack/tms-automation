import unittest
import os
import sys

# Windows 콘솔에서 유니코드 안전하게 찍기(UTF-8)
try:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    # Logger가 sys.__stdout__에 직접 쓰므로 원본도 보강
    if hasattr(sys, "__stdout__") and hasattr(sys.__stdout__, "reconfigure"):
        sys.__stdout__.reconfigure(encoding="utf-8", errors="replace")
        sys.__stderr__.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

import time
from datetime import datetime

import pandas as pd
import allure

from appium import webdriver
from appium.options.android import UiAutomator2Options
from appium.webdriver.common.appiumby import AppiumBy

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait   # ✅ 추가
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.common.actions.pointer_input import PointerInput
from selenium.common.exceptions import InvalidSessionIdException, WebDriverException

from openpyxl import load_workbook


# ===== 유틸: Config/Accounts/변수치환/마스킹 =====
import re, json, math, collections


def get_cfg(cfg: dict, key: str, default=None):
    val = cfg.get(key, default)
    if isinstance(val, float) and pd.isna(val):
        return default
    return val

def load_accounts(xlsx_path: str, sheet_name: str = "Accounts"):
    """
    Accounts 시트에서 active=Y 인 계정만 [{'ID':..., 'PWD':...}, ...] 로드
    """
    try:
        df = pd.read_excel(xlsx_path, sheet_name=sheet_name)
    except Exception:
        return []
    def norm(v): 
        return "" if pd.isna(v) else str(v).strip()
    rows = []
    for _, r in df.iterrows():
        active = norm(r.get("active", "Y")).upper()
        if active not in ("Y","YES","TRUE","1"):
            continue
        rid = norm(r.get("id", ""))
        pwd = norm(r.get("password", ""))
        if rid and pwd:
            rows.append({"ID": rid, "PWD": pwd})
    return rows

VAR_RE = re.compile(r"\$\{([A-Za-z0-9_]+)\}")

def substitute_vars(s: str, context: dict):
    """
    문자열 s에서 ${KEY} 를 context[KEY] 로 치환. (KEY 없으면 원문 유지)
    """
    if not s or not isinstance(s, str):
        return s
    def repl(m):
        key = m.group(1)
        return str(context.get(key, m.group(0)))
    return VAR_RE.sub(repl, s)

def mask_secret(txt: str, enabled=True):
    """
    로그/리포트에 노출될 민감 문자열 마스킹
    """
    if not enabled or not txt:
        return txt
    return "*" * min(len(txt), 8)

def safe_filename(name, fallback="screenshot"):
        import math, re
        # None/NaN/'nan'/공란 → fallback
        try:
            if name is None:
                s = fallback
            elif isinstance(name, float) and math.isnan(name):
                s = fallback
            else:
                s = str(name)
        except Exception:
            s = fallback

        s = s.strip()
        if not s or s.lower() == "nan":
            s = fallback
        return re.sub(r'[\/:*?"<>|]', '_', s)



# --- KEY INPUT MAP (문자 -> (keycode, needs_shift)) ---
# 참고: A..Z = 29..54, 0..9 = 7..16
KEYCODE_MAP = {
    # lower
    "a": (29, False), "b": (30, False), "c": (31, False), "d": (32, False),
    "e": (33, False), "f": (34, False), "g": (35, False), "h": (36, False),
    "i": (37, False), "j": (38, False), "k": (39, False), "l": (40, False),
    "m": (41, False), "n": (42, False), "o": (43, False), "p": (44, False),
    "q": (45, False), "r": (46, False), "s": (47, False), "t": (48, False),
    "u": (49, False), "v": (50, False), "w": (51, False), "x": (52, False),
    "y": (53, False), "z": (54, False),

    # upper (SHIFT 조합)
    "A": (29, True),  "B": (30, True),  "C": (31, True),  "D": (32, True),
    "E": (33, True),  "F": (34, True),  "G": (35, True),  "H": (36, True),
    "I": (37, True),  "J": (38, True),  "K": (39, True),  "L": (40, True),
    "M": (41, True),  "N": (42, True),  "O": (43, True),  "P": (44, True),
    "Q": (45, True),  "R": (46, True),  "S": (47, True),  "T": (48, True),
    "U": (49, True),  "V": (50, True),  "W": (51, True),  "X": (52, True),
    "Y": (53, True),  "Z": (54, True),

    # digits
    "0": (7, False), "1": (8, False), "2": (9, False),  "3": (10, False),
    "4": (11, False), "5": (12, False), "6": (13, False), "7": (14, False),
    "8": (15, False), "9": (16, False),

    # whitespace / control
    " ": (62, False),     # SPACE
    "\n": (66, False),    # ENTER
    "\r": (66, False),    # ENTER
    "\t": (61, False),    # TAB
    "\b": (67, False),    # DEL(BACKSPACE)

    # punctuation (일부는 SHIFT 필요)
    ",": (55, False),
    "<": (55, True),      # SHIFT + COMMA
    ".": (56, False),
    ">": (56, True),      # SHIFT + PERIOD
    "/": (76, False),
    "?": (76, True),      # SHIFT + SLASH
    ";": (74, False),
    ":": (74, True),      # SHIFT + SEMICOLON
    "'": (75, False),
    "\"": (75, True),     # SHIFT + APOSTROPHE
    "[": (71, False),
    "{": (71, True),      # SHIFT + LEFT_BRACKET
    "]": (72, False),
    "}": (72, True),      # SHIFT + RIGHT_BRACKET
    "\\": (73, False),
    "|": (73, True),      # SHIFT + BACKSLASH
    "-": (69, False),
    "_": (69, True),      # SHIFT + MINUS
    "=": (70, False),
    "+": (70, True),      # SHIFT + EQUALS
    "@": (77, False),     # AT
    "#": (18, True),      # 보통 SHIFT+3 (단말/IME에 따라 다를 수 있음)
    "!": (8,  True),      # SHIFT+1
    "(": (10, True),      # SHIFT+9
    ")": (11, True),      # SHIFT+0
}

# SHIFT 메타 상태 (Android KeyEvent META_SHIFT_ON)
META_SHIFT_ON = 0x00000001


class Logger(object):
    def __init__(self):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.filename = f"logs/test_run_{timestamp}.log"
        self.terminal = sys.stdout # 파이프를 통해 한 번만 흘려보내기
        # 줄 단위 버퍼링 + UTF-8
        self.log = open(self.filename, "w", encoding="utf-8", buffering=1, newline="")

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)
        # 줄바꿈 포함될 때만 flush → 대폭 빨라짐
        if "\n" in message:
            self.flush()

    def flush(self):
        try:
            self.terminal.flush()
        except Exception:
            pass
        try:
            self.log.flush()
        except Exception:
            pass
        

def log_debug(msg): print(f"\033[90m[DEBUG] {msg}\033[0m")
def log_info(msg): print(f"\033[96m[INFO] {msg}\033[0m")
def log_warn(msg): print(f"\033[93m[WARN] {msg}\033[0m")
def log_ok(msg): print(f"\033[92m[OK] {msg}\033[0m")
def log_fail(msg): print(f"\033[91m[FAIL] {msg}\033[0m")
def log_skip(msg): print(f"\033[95m[SKIP] {msg}\033[0m")

class ExcelDrivenAppiumTest(unittest.TestCase):
    # NaN/빈문자/None 안전 변환기
    def _cfg_num(self, cfg: dict, key: str, default, as_int: bool = True):
        val = cfg.get(key, default)
        try:
            # None, "", "nan", NaN → 기본값
            if val is None:
                return default
            s = str(val).strip()
            if s == "" or s.lower() == "nan":
                return default
            f = float(s)
            if math.isnan(f) or math.isinf(f):
                return default
            return int(f) if as_int else float(f)
        except Exception:
            return default

    def setUp(self):
        options = UiAutomator2Options()
        options.set_capability("udid", "R3CMB0E09XM")
        options.set_capability("platformName", "Android")
        options.set_capability("automationName", "UiAutomator2")
        options.set_capability("appPackage", "com.tms")
        options.set_capability("noReset", True)
        self.driver = webdriver.Remote("http://localhost:4723", options=options)
        self.wait = WebDriverWait(self.driver, 20)
        self.action_file = "actions_tms_updated.xlsx"
        os.makedirs("allure-results/screenshots", exist_ok=True)
        os.makedirs("logs", exist_ok=True)
        self._logger = Logger()
        sys.stdout = self._logger
        sys.stderr = self._logger


    def add_screenshot(self, name=None):
        """
        현재 화면을 캡처해서 allure-results/screenshots 폴더에 저장하고,
        Allure 리포트에 첨부.  # NEW: 성공 시 파일 경로를 반환
        """
        import os, math, allure

        # 렌더링 안정화용 추가 대기 (옵션)
        self._sleep_ms(getattr(self, "screenshot_extra_wait_ms", 0))


        # 이름 보정 (None, NaN, 빈 문자열, 'nan' 모두 fallback 처리)
        label = name
        try:
            if label is None or (isinstance(label, float) and math.isnan(label)) \
            or str(label).strip().lower() in ("", "nan"):
                label = "screenshot"
        except Exception:
            label = "screenshot"

        filename = safe_filename(label)
        screenshot_dir = os.path.join("allure-results", "screenshots")
        os.makedirs(screenshot_dir, exist_ok=True)
        filepath = os.path.join(screenshot_dir, f"{filename}.png")

        try:
            self.driver.get_screenshot_as_file(filepath)
            allure.attach.file(filepath, name=label, attachment_type=allure.attachment_type.PNG)
            log_ok(f"Screenshot saved: {filepath}")
            return filepath  # ✅ NEW
        except Exception as e:
            log_warn(f"[WARN] Screenshot capture failed: {e}")
            return None      # ✅ NEW

    def _sleep_ms(self, ms: int):
        import time
        try:
            if ms and int(ms) > 0:
                time.sleep(float(ms) / 1000.0)
        except Exception:
            pass

    def _capture_stable_page_source(self, tries: int = 4, interval_ms: int = 200):
        """
        page_source가 연속 2회 동일해질 때까지 재시도하여 안정화된 XML을 반환.
        반환: (xml_text:str, is_stable:bool)
        """
        import hashlib, time
        prev = None
        prev_md5 = None
        for i in range(max(1, int(tries))):
            src = self.driver.page_source or ""
            md5 = hashlib.md5(src.encode("utf-8", "ignore")).hexdigest()
            if prev_md5 and md5 == prev_md5:
                return src, True
            prev = src
            prev_md5 = md5
            time.sleep(float(interval_ms) / 1000.0)
        return prev or "", False



    def _save_baseline(self, step_key, img_path, page_src, xpath):
        import os, json, shutil
        base_dir = self.snapshot_dir or "snapshots"
        img_dir  = os.path.join(base_dir, "img");  os.makedirs(img_dir, exist_ok=True)
        xml_dir  = os.path.join(base_dir, "xml");  os.makedirs(xml_dir, exist_ok=True)
        meta_dir = os.path.join(base_dir, "meta"); os.makedirs(meta_dir, exist_ok=True)

        # 이미지 복사
        try:
            if img_path and os.path.exists(img_path):
                shutil.copyfile(img_path, os.path.join(img_dir, f"{step_key}.png"))
        except Exception as e:
            log_warn(f"[BASELINE] 이미지 저장 경고: {e}")

        # page_source 저장
        try:
            with open(os.path.join(xml_dir, f"{step_key}.xml"), "w", encoding="utf-8") as f:
                f.write(page_src or "")
        except Exception as e:
            log_warn(f"[BASELINE] XML 저장 경고: {e}")

        # 메타 저장 (주요 앵커/XPATH, 액티비티)
        meta = {"xpath": xpath, "activity": getattr(self.driver, "current_activity", None)}
        try:
            with open(os.path.join(meta_dir, f"{step_key}.json"), "w", encoding="utf-8") as f:
                json.dump(meta, f, ensure_ascii=False, indent=2)
        except Exception as e:
            log_warn(f"[BASELINE] 메타 저장 경고: {e}")

        log_ok(f"[BASELINE] saved: {step_key}")


    def _verify_against_baseline(self, step_key, img_path, page_src, xpath_now, action=None):
        import os, json, hashlib
        from selenium.webdriver.common.by import By

        base_dir = self.snapshot_dir or "snapshots"
        img_file  = os.path.join(base_dir, "img",  f"{step_key}.png")
        xml_file  = os.path.join(base_dir, "xml",  f"{step_key}.xml")
        meta_file = os.path.join(base_dir, "meta", f"{step_key}.json")

        def _md5_bytes(b: bytes):
            return hashlib.md5(b).hexdigest() if b else None
        def _md5_file(p: str):
            if not (p and os.path.exists(p)): return None
            with open(p, "rb") as f:
                return hashlib.md5(f.read()).hexdigest()

        policy = (self.verify_policy or "WARN").upper()

        # 1) XPATH 재검출: click/tap 이후엔 post-검사 스킵
        if os.path.exists(meta_file):
            with open(meta_file, "r", encoding="utf-8") as f:
                meta = json.load(f)
            xp = meta.get("xpath") or xpath_now
            if xp:
                act = (str(action).lower() if action is not None else "")
                if act in ("click", "tap"):
                    log_info("[VERIFY] click/tap 이후 step → post-XPATH 체크 스킵")
                else:
                    if not self.driver.find_elements(By.XPATH, xp):
                        raise AssertionError(f"[VERIFY] XPATH 미검출: {xp}")

        # 2) XML 비교 (완전 동일성) — 정책 적용
        base_xml = open(xml_file, "rb").read() if os.path.exists(xml_file) else b""
        cur_xml  = (page_src or "").encode("utf-8", "ignore")
        if _md5_bytes(base_xml) and _md5_bytes(cur_xml) and _md5_bytes(base_xml) != _md5_bytes(cur_xml):
            msg = "[VERIFY] XML 차이 감지"
            if policy in ("STRICT", "STRICT_XML"):
                raise AssertionError(msg)
            else:
                log_warn(msg)

        # 3) 이미지 비교 — 전략/민감도/크롭 적용
        strategy   = str(getattr(self, "verify_img_strategy", "EXACT") or "EXACT").upper()
        thresh_pct = float(getattr(self, "verify_img_max_diff_pct", 0.8))
        ssim_thr   = float(getattr(self, "verify_img_ssim_threshold", 0.98))

        crop_t = int(getattr(self, "img_ignore_top_px", 0))
        crop_b = int(getattr(self, "img_ignore_bottom_px", 0))
        crop_l = int(getattr(self, "img_ignore_left_px", 0))
        crop_r = int(getattr(self, "img_ignore_right_px", 0))

        same = False
        try:
            from PIL import Image
            import numpy as np

            def _load_crop(p):
                im = Image.open(p).convert("RGB")
                w, h = im.size
                L = min(max(crop_l, 0), w-1)
                T = min(max(crop_t, 0), h-1)
                R = max(w - max(crop_r, 0), L+1)
                B = max(h - max(crop_b, 0), T+1)
                im = im.crop((L, T, R, B))
                return np.array(im)

            if os.path.exists(img_file) and os.path.exists(img_path):
                a = _load_crop(img_file)
                b = _load_crop(img_path)
                if a.shape != b.shape:
                    b = Image.fromarray(b).resize((a.shape[1], a.shape[0]))
                    b = np.array(b)

                if strategy == "EXACT":
                    same = np.array_equal(a, b)

                elif strategy == "DIFF":
                    # 채널별 오차 허용치(±3/255)로 다른 픽셀 카운트 → 퍼센트
                    diff = np.abs(a.astype(np.int16) - b.astype(np.int16))
                    diff_pix = np.any(diff > 3, axis=2).sum()
                    total = a.shape[0] * a.shape[1]
                    diff_pct = (diff_pix / total) * 100.0
                    log_info(f"[VERIFY][IMG] DIFF≈{diff_pct:.3f}% (thr {thresh_pct}%)")
                    same = (diff_pct <= thresh_pct)

                elif strategy == "SSIM":
                    # 간단한 그레이 SSIM 근사
                    ag = (0.299*a[:,:,0] + 0.587*a[:,:,1] + 0.114*a[:,:,2]).astype(np.float32)
                    bg = (0.299*b[:,:,0] + 0.587*b[:,:,1] + 0.114*b[:,:,2]).astype(np.float32)
                    mu_a = ag.mean(); mu_b = bg.mean()
                    sigma_a = ag.var(); sigma_b = bg.var()
                    sigma_ab = ((ag - mu_a)*(bg - mu_b)).mean()
                    C1 = 6.5025; C2 = 58.5225
                    ssim = ((2*mu_a*mu_b + C1)*(2*sigma_ab + C2)) / ((mu_a**2 + mu_b**2 + C1)*(sigma_a + sigma_b + C2))
                    log_info(f"[VERIFY][IMG] SSIM≈{ssim:.4f} (thr {ssim_thr})")
                    same = (ssim >= ssim_thr)

                else:
                    # 알 수 없는 전략 → EXACT로 처리
                    same = np.array_equal(a, b)

        except Exception as e:
            log_warn(f"[VERIFY][IMG] 고급 비교 실패({e}) → MD5 fallback")
            base_png_md5 = _md5_file(img_file)
            cur_png_md5  = _md5_file(img_path)
            same = (base_png_md5 and cur_png_md5 and base_png_md5 == cur_png_md5)

        if not same:
            msg = "[VERIFY] 스크린샷 차이 감지"
            if policy in ("STRICT", "STRICT_IMG"):
                raise AssertionError(msg)
            else:
                log_warn(msg)


        log_ok(f"[VERIFY] passed: {step_key}")




    def type_keys(self, text: str, delay: float = 0.02):
        unknown = []
        for ch in str(text):
            info = KEYCODE_MAP.get(ch)
            if info is None:
                unknown.append(ch)
                continue
            keycode, need_shift = info
            if need_shift:
                # SHIFT down → key → SHIFT up
                # Appium Python의 press_keycode는 metastate를 함께 보낼 수 있음
                self.driver.press_keycode(keycode, META_SHIFT_ON)
            else:
                self.driver.press_keycode(keycode)
            time.sleep(delay)

        if unknown:
            # 매핑되지 않은 문자는 유니코드 친화적인 fallback으로 입력
            try:
                self.driver.execute_script("mobile: type", {"text": "".join(unknown)})
            except Exception:
                # 최후수단: 클립보드 붙여넣기(가능한 경우)
                try:
                    self.driver.set_clipboard_text("".join(unknown))
                    # KEYCODE_PASTE(279)는 일부 단말/IME에서만 동작
                    self.driver.press_keycode(279)
                except Exception:
                    # 그래도 실패하면 로그만 남김
                    print(f"[WARN] unmapped chars not typed: {''.join(unknown)}")


    def perform_action(
        self,
        action,
        by,
        value,
        name="",
        sleep_time=0,
        skip_on_error="N",
        visible_if_type="text",
        visible_if="",
        mandatory="Y",
        no="",
        input_text=""
    ):
        try:
            # ---- 입력 정규화 ----
            visible_if = "" if (pd.isna(visible_if)) else str(visible_if)
            visible_if_type = str(visible_if_type).strip().lower() if not pd.isna(visible_if_type) else "text"
            skip_on_error = str(skip_on_error).strip().upper() if not pd.isna(skip_on_error) else "N"
            mandatory = str(mandatory).strip().upper() if not pd.isna(mandatory) else "Y"
            input_text = "" if pd.isna(input_text) else str(input_text)

            # ---- 로그 헤더 ----
            print("\n" + "=" * 70)
            print(f"[STEP {no:>3}] >> '{name}' 실행 시작")
            print(f"{'- Action'.ljust(20)}: {action}")
            print(f"{'- By / Value'.ljust(20)}: {by} -> {value}")
            print(f"{'- visible_if'.ljust(20)}: '{visible_if}' (type: {visible_if_type})")
            print(f"{'- skip_on_error'.ljust(20)}: {skip_on_error}")
            print(f"{'- mandatory'.ljust(20)}: {mandatory}")
            if action == "input":
                print(f"{'- input_text'.ljust(20)}: {input_text}")

            # ---- 표시 조건(가드) ----
            if visible_if:
                try:
                    if visible_if_type == "text":
                        if visible_if not in self.driver.page_source:
                            log_skip(f"텍스트 '{visible_if}' 없음 → Step({no}) SKIP")
                            if mandatory == "Y" and skip_on_error != "Y":
                                raise AssertionError(f"Required step failed (visible_if text not found): {visible_if}")
                            return False
                    elif visible_if_type == "xpath":
                        if not self.driver.find_elements(By.XPATH, visible_if):
                            log_skip(f"요소 '{visible_if}' 없음 → Step({no}) SKIP")
                            if mandatory == "Y" and skip_on_error != "Y":
                                raise AssertionError(f"Required step failed (visible_if xpath not found): {visible_if}")
                            return False
                    else:
                        log_warn(f"알 수 없는 visible_if_type='{visible_if_type}' → 무시")
                except Exception as ve:
                    log_warn(f"visible_if 평가 오류: {ve} → Step({no}) SKIP")
                    if mandatory == "Y" and skip_on_error != "Y":
                        raise AssertionError(f"Required step failed (visible_if evaluation error): {ve}")
                    return False

            # ---- 대기 ----
            if sleep_time and not pd.isna(sleep_time):
                log_info(f"WAIT {sleep_time} seconds")
                time.sleep(float(sleep_time))

            # ---- 스크린샷(사전) ----
            self.add_screenshot(name or value)

            # ✅ PRE-XPATH 검증: VERIFY 모드 + click/input + XPATH일 때
            if getattr(self, "snapshot_mode", "OFF") == "VERIFY":
                try:
                    act = str(action).lower()
                except Exception:
                    act = ""
                by_u = str(by).upper() if not pd.isna(by) else ""
                if act in ("click", "input") and by_u == "XPATH":
                    val = "" if pd.isna(value) else str(value)
                    if not val:
                        raise AssertionError("[VERIFY:PRE] XPATH 값이 비어 있습니다.")
                    if not self.driver.find_elements(By.XPATH, val):
                        raise AssertionError(f"[VERIFY:PRE] XPATH 미검출: {val}")

            # ==========================
            #        액션 분기
            # ==========================

            # (기존 click/input/tap/swipe/key/back 분기 그대로… — 생략 없음)
            # 1) 요소 기반: click / input
            if action in ("click", "input"):
                by_u = str(by).upper() if not pd.isna(by) else ""
                val = "" if pd.isna(value) else str(value)

                # By 매핑 (selenium.By, appium.AppiumBy 둘 다 시도)
                locator = getattr(By, by_u, None) or getattr(AppiumBy, by_u, None)
                if not locator:
                    raise ValueError(f"[{action}] 지원하지 않는 By: {by}")

                # 요소 찾기
                try:
                    elements = self.driver.find_elements(locator, val)
                except Exception:
                    elements = []
                if not elements:
                    log_skip(f"Element not present: {val} → Step({no}) SKIP")
                    if mandatory == "Y" and skip_on_error != "Y":
                        raise AssertionError(f"Required step failed (element not present): {val}")
                    return False

                el = elements[0]
                if action == "click":
                    # 기본 click → 실패 시 중앙 좌표 TAP으로 폴백
                    try:
                        el.click()
                    except Exception as e:
                        log_warn(f"[click] el.click() 실패 → 좌표 TAP 폴백: {e}")
                        rect = el.rect
                        cx = int(rect["x"] + rect["width"] / 2)
                        cy = int(rect["y"] + rect["height"] / 2)
                        finger = PointerInput("touch", "finger1")
                        actions = ActionBuilder(self.driver, mouse=finger)
                        actions.pointer_action.move_to_location(cx, cy)
                        actions.pointer_action.pointer_down()
                        actions.pointer_action.pause(0.05)
                        actions.pointer_action.pointer_up()
                        actions.perform()
                    log_ok(f"Clicked: {val}")
                else:
                    # input
                    el.clear()
                    el.send_keys(input_text)
                    log_ok(f"Input: {mask_secret(input_text, True)}")

            # 2) 좌표 TAP (coord 또는 abs)
            elif action == "tap":
                value_str = "" if pd.isna(value) else str(value)
                if not value_str or "," not in value_str:
                    raise ValueError(f"[TAP] 유효하지 않은 좌표 입력: '{value_str}'")

                by_lower = str(by).lower() if by is not None else ""
                if by_lower == "coord":
                    xr, yr = map(float, value_str.split(","))
                    size = self.driver.get_window_size()
                    x = int(size["width"] * xr)
                    y = int(size["height"] * yr)
                elif by_lower == "abs":
                    x, y = map(int, value_str.split(","))
                else:
                    raise ValueError(f"[TAP] Unsupported tap type: {by}")

                finger = PointerInput("touch", "finger1")
                last_err = None
                for attempt in range(1, 4):
                    try:
                        actions = ActionBuilder(self.driver, mouse=finger)
                        actions.pointer_action.move_to_location(x, y)
                        actions.pointer_action.pointer_down()
                        actions.pointer_action.pause(0.05)
                        actions.pointer_action.pointer_up()
                        actions.perform()
                        log_ok(f"W3C Tapped: ({x},{y}) [시도 {attempt}]")
                        break
                    except Exception as e:
                        last_err = e
                        log_warn(f"[TAP 실패] ({x},{y}) 시도 {attempt} → {e}")
                        time.sleep(1.0)
                else:
                    raise ValueError(f"[TAP] 좌표 TAP 실패 (3회 시도) → {last_err}")

            # 3) 키 입력 (키코드 시뮬레이션)
            elif action == "key":
                self.type_keys(input_text)
                log_ok(f"Key 입력: {mask_secret(input_text, True)}")

            # 4) 뒤로가기
            elif action == "back":
                self.driver.back()
                log_ok("Back pressed")

            # 3) 스와이프 (x1,y1,x2,y2[,duration_ms])
            elif action == "swipe":
                value_str = "" if pd.isna(value) else str(value)
                if not value_str or "," not in value_str:
                    raise ValueError(f"[SWIPE] 유효하지 않은 좌표 입력: '{value_str}'")
                try:
                    parts = [p.strip() for p in value_str.replace('|', ',').replace(';', ',').split(',')]
                    if len(parts) < 4:
                        raise ValueError(f"[SWIPE] 좌표 부족: '{value_str}'")

                    x1, y1, x2, y2 = [int(float(n)) for n in parts[:4]]
                    duration_ms = int(float(parts[4])) if len(parts) >= 5 and parts[4] else 800
                    duration_ms = max(100, duration_ms)

                    self.driver.swipe(x1, y1, x2, y2, duration_ms)
                    log_ok(f"Swiped from ({x1},{y1}) to ({x2},{y2}) in {duration_ms} ms")
                except Exception as se:
                    raise ValueError(f"[SWIPE] 좌표 파싱/실행 오류: '{value_str}' → {se}")

            # --- 스텝 완료 로그 ---
            log_ok(f"Step [{no}] '{name}' 완료")

            # --- Snapshot: after-action capture & compare ---
            if getattr(self, "snapshot_mode", "OFF") in ("BASELINE", "VERIFY"):
                step_key = f"{self.current_sheet}__{str(no)}__{safe_filename(name or str(value) or 'step')}"
                post_img_path = self.add_screenshot(f"{name}_post")
                page_src = self.driver.page_source
                # 액션 이후 캡처 안정화
                self._sleep_ms(getattr(self, "post_capture_wait_ms", 300))  # 1) 고정 딜레이
                page_src, stable = self._capture_stable_page_source(
                    tries=getattr(self, "post_capture_retry", 4),
                    interval_ms=getattr(self, "post_capture_interval_ms", 200),
                )  # 2) DOM 안정화
                post_img_path = self.add_screenshot(f"{name}_post")         # 3) 스크린샷은 마지막에
                if not stable:
                    log_warn("[VERIFY] page_source 안정화 실패(재시도 한계) → 마지막 스냅샷으로 비교 진행")
                locator_xpath = value if (str(by).upper() == "XPATH") else ""
                if self.snapshot_mode == "BASELINE":
                    self._save_baseline(step_key, post_img_path, page_src, locator_xpath)
                else:  # VERIFY
                    # ✅ action 인자 추가 전달
                    self._verify_against_baseline(step_key, post_img_path, page_src, locator_xpath, action)

            return True

        except (InvalidSessionIdException, WebDriverException) as e:
            if "terminated" in str(e) or "not started" in str(e):
                log_fail(f"[세션 종료] Step [{no}] '{name}' - Appium 세션 종료")
            raise
        except Exception as e:
            log_fail(f"Step [{no}] '{name}' - {str(e)}")
            self.add_screenshot(f"{name}_error")
            if skip_on_error == "Y" or mandatory == "N":
                log_warn("Continue after error due to skip_on_error=Y or mandatory=N")
                return False
            raise



    def run_sheet_range_with_context(self, sheet_name: str, start_no: int, end_no: int, context: dict, cfg: dict):

    # 시트에서 no ∈ [start_no, end_no] 범위만 실행.
    # 실행 시 ${ID}/${PWD} 같은 플레이스홀더를 context 값으로 치환.

        df = pd.read_excel(self.action_file, sheet_name=sheet_name)
        df = df[(df["no"] >= start_no) & (df["no"] <= end_no)].sort_values(by="no")

        mask_on = str(get_cfg(cfg, "mask_secrets_in_logs", "Y")).upper() in ("Y","YES","TRUE","1")

        for _, row in df.iterrows():
            action     = str(row.get("action","")).lower()
            by         = row.get("by","")
            value      = row.get("value","")
            input_text = row.get("input_text","")

            name_raw = row.get("name", "")
            try:
                name_str = "" if pd.isna(name_raw) else str(name_raw).strip()
            except Exception:
                name_str = ""
            if not name_str or name_str.lower() == "nan":
                name_str = f"step_{row.get('no','')}"

            value_str       = "" if pd.isna(value) else str(value)
            input_text_raw  = "" if pd.isna(input_text) else str(input_text)
            value_resolved  = substitute_vars(value_str, context or {})
            input_resolved  = substitute_vars(input_text_raw, context or {})

            # (선택) 여기서 프리뷰 로그를 쓰려면 mask_secret(input_resolved, mask_on) 사용 가능
            # preview = input_resolved if action != "key" else mask_secret(input_resolved, mask_on)

            self.perform_action(
                action=action,
                by=by,
                value=value_resolved,
                name=name_str,
                sleep_time=row.get("sleep",0),
                skip_on_error=row.get("skip_on_error","N"),
                visible_if_type=row.get("visible_if_type", row.get("jump_if_visible_type", "text")),
                visible_if=row.get("visible_if", row.get("jump_if_visible","")),
                mandatory=row.get("mandatory","Y"),
                no=row.get("no",""),
                input_text=input_resolved
            )

    def run_sheet_with_context(self, sheet_name: str, cfg: dict, context: dict = None):

        # 시트를 처음부터 끝까지 실행. (기존 단일 실행용)
        # 플레이스홀더가 있으면 context로 치환, 없으면 원문 그대로.

        df = pd.read_excel(self.action_file, sheet_name=sheet_name).sort_values(by="no")
        mask_on = str(get_cfg(cfg, "mask_secrets_in_logs", "Y")).upper() in ("Y","YES","TRUE","1")

        for _, row in df.iterrows():
            action     = str(row.get("action","")).lower()
            by         = row.get("by","")
            value      = row.get("value","")
            input_text = row.get("input_text","")

            name_raw = row.get("name", "")
            try:
                name_str = "" if pd.isna(name_raw) else str(name_raw).strip()
            except Exception:
                name_str = ""
            if not name_str or name_str.lower() == "nan":
                name_str = f"step_{row.get('no','')}"

            value_str       = "" if pd.isna(value) else str(value)
            input_text_raw  = "" if pd.isna(input_text) else str(input_text)
            value_resolved  = substitute_vars(value_str, context or {})
            input_resolved  = substitute_vars(input_text_raw, context or {})

            self.perform_action(
                action=action,
                by=by,
                value=value_resolved,
                name=name_str,
                sleep_time=row.get("sleep",0),
                skip_on_error=row.get("skip_on_error","N"),
                visible_if_type=row.get("visible_if_type", row.get("jump_if_visible_type", "text")),
                visible_if=row.get("visible_if", row.get("jump_if_visible","")),
                mandatory=row.get("mandatory","Y"),
                no=row.get("no",""),
                input_text=input_resolved
            )



    @allure.feature("Excel Driven UI Test")
    def test_actions_from_excel(self):
        all_sheets = pd.read_excel(self.action_file, sheet_name=None)
        config_df = all_sheets.get("Config")
        if config_df is None or "run_sheet_name" not in config_df.columns:
            log_fail("❌ Config 시트가 없거나 'run_sheet_name' 열이 없습니다.")
            return

        # Config 행들을 순회(각 행마다 다른 범위/옵션을 줄 수 있음)
        for idx, row in config_df.iterrows():
            # NEW: enabled 플래그 해석 (기본 Y)
            enabled_raw = str(row.get("enabled", "Y")).strip().upper()
            is_enabled = enabled_raw in ("Y", "YES", "TRUE", "1")
            if not is_enabled:
                log_skip(f"[SKIP] enabled={enabled_raw} → 시나리오 미실행 (row={idx+2})")
                continue

            run_sheet = row.get("run_sheet_name")
            if not pd.notna(run_sheet) or run_sheet not in all_sheets:
                log_warn(f"[SKIP] 유효하지 않은 시트명: {run_sheet}")
                continue

            log_info(f"\n📘 실행할 시트: '{run_sheet}'")
            # ----- loop 옵션 해석 -----
            loop_accounts = str(row.get("loop_accounts", "N")).upper() in ("Y","YES","TRUE","1")
            accounts_sheet_name = row.get("accounts_sheet_name", "Accounts")
            account_loop_scope  = str(row.get("account_loop_scope", "3~16")).strip()
            restart_each        = str(row.get("restart_app_each_account", "N")).upper() in ("Y","YES","TRUE","1")

            # Config를 dict로 (도우미에서 사용)
            cfg = { col: row.get(col) for col in config_df.columns }

            # Snapshot 옵션 (엑셀 Config에서 읽음)
            self.snapshot_mode = str(cfg.get("snapshot_mode", "OFF")).strip().upper()
            self.snapshot_dir  = str(cfg.get("snapshot_dir", "snapshots")).strip()
            self.verify_policy = str(cfg.get("verify_policy", "WARN")).strip().upper()
            self.current_sheet = str(row.get("run_sheet_name", "")).strip()

            # 캡처 안정화 옵션 (빈칸/NaN 안전)
            self.post_capture_wait_ms     = self._cfg_num(cfg, "post_capture_wait_ms", 300, as_int=True)
            self.post_capture_retry       = self._cfg_num(cfg, "post_capture_retry", 4, as_int=True)
            self.post_capture_interval_ms = self._cfg_num(cfg, "post_capture_interval_ms", 200, as_int=True)
            self.screenshot_extra_wait_ms = self._cfg_num(cfg, "screenshot_extra_wait_ms", 0, as_int=True)

            # === 이미지 비교 민감도 옵션 ===
            self.verify_img_strategy       = str(cfg.get("verify_img_strategy", "EXACT")).strip().upper()  # EXACT | DIFF | SSIM
            self.verify_img_max_diff_pct   = self._cfg_num(cfg, "verify_img_max_diff_pct", 0.8, as_int=False)
            self.verify_img_ssim_threshold = self._cfg_num(cfg, "verify_img_ssim_threshold", 0.98, as_int=False)

            # 배너/상태바/네비바 등 무시할 크롭 픽셀(양끝 방향별) — 빈칸/NaN 허용
            self.img_ignore_top_px    = self._cfg_num(cfg, "img_ignore_top_px", 0, as_int=True)
            self.img_ignore_bottom_px = self._cfg_num(cfg, "img_ignore_bottom_px", 0, as_int=True)
            self.img_ignore_left_px   = self._cfg_num(cfg, "img_ignore_left_px", 0, as_int=True)
            self.img_ignore_right_px  = self._cfg_num(cfg, "img_ignore_right_px", 0, as_int=True)



            # ----- 계정 루프 모드 -----
            if loop_accounts:
                accounts = load_accounts(self.action_file, sheet_name=accounts_sheet_name if pd.notna(accounts_sheet_name) else "Accounts")
                if not accounts:
                    log_warn("[WARN] 활성 계정 없음/Accounts 시트 미존재 → 기존 단일 실행으로 전환합니다.")
                    self.run_sheet_with_context(run_sheet, cfg, context=None)
                    continue

                # 범위 파싱 (예: "3~16")
                try:
                    start_no, end_no = map(int, account_loop_scope.replace(" ", "").split("~"))
                except Exception:
                    start_no, end_no = 3, 16
                    log_warn(f"[WARN] account_loop_scope 파싱 실패 → 기본값 3~16 사용")

                for a_idx, acc in enumerate(accounts, 1):
                    log_info(f"\n===== 계정 루프 {a_idx}/{len(accounts)}: {acc['ID']} =====")

                    if restart_each and a_idx > 1:
                        # 계정별 앱 재기동(옵션)
                        try:
                            self.driver.quit()
                        except Exception as e:
                            log_warn(f"[restart] driver.quit() 실패: {e}")
                        # 세션 재생성
                        self.setUp()

                    # no ∈ [start_no, end_no]만 실행 (ID/PWD 치환)
                    self.run_sheet_range_with_context(run_sheet, start_no, end_no, acc, cfg)

                # (필요시) 루프 이후 후속 step 실행 예:
                # self.run_sheet_range_with_context(run_sheet, end_no+1, 9999, context=None, cfg=cfg)
                # 지금은 요구사항에 없으니 생략

                # 계정 루프 모드에서는 last_failed_step 초기화만 수행
                self.update_config_cell(idx, "last_failed_step", "")
                continue

            # ----- 단일 실행 모드 (기존 로직 유지) -----
            df = all_sheets[run_sheet].copy()
            df["no_str"] = df["no"].apply(lambda x: str(int(float(x))) if pd.notna(x) else "")
            df_steps = df.copy()

            scope_raw = str(row.get("test_step_scope", "")).strip()
            if scope_raw and scope_raw != "nan":
                step_list = []
                for part in scope_raw.split(','):
                    part = part.strip()
                    if '~' in part:
                        a, b = map(int, part.split('~'))
                        step_list.extend(range(a, b + 1))
                    else:
                        step_list.append(int(part))
                df_steps = df[df["no"].isin(step_list)]
                log_warn(f"[범위설정] test_step_scope 사용됨 → {scope_raw} → 총 {len(df_steps)}개 step 실행")
            else:
                # 기존 index 기반 범위/재시작
                start_index = 0
                end_index = len(df)
                last_failed_no = row.get("last_failed_step", "")
                start_no = row.get("start_step_no", "")
                end_no   = row.get("end_step_no", "")
                if pd.notna(last_failed_no):
                    match = df[df["no_str"] == str(int(float(last_failed_no)))].index
                    if not match.empty:
                        start_index = match[0]
                        log_warn(f"[재시작] last_failed_step → {last_failed_no} (index={start_index})")
                elif pd.notna(start_no):
                    match = df[df["no_str"] == str(int(float(start_no)))].index
                    if not match.empty:
                        start_index = match[0]
                        log_warn(f"[시작] start_step_no → {start_no} (index={start_index})")
                if pd.notna(end_no):
                    match = df[df["no_str"] == str(int(float(end_no)))].index
                    if not match.empty:
                        end_index = match[0] + 1
                        log_warn(f"[범위제한] end_step_no → {end_no} (index={end_index - 1})")

                df_steps = df.iloc[start_index:end_index]

            # 단일 실행 루프 (포인터 방식: for 제거, while+k만 사용)
            indexes = list(df_steps.index)
            k = 0
            visited = collections.defaultdict(int)

            while k < len(indexes):
                index = indexes[k]
                row_step = df.loc[index]
                visited[index] += 1
                if visited[index] > 5:
                    log_warn(f"무한루프 방지: Step index {index}가 5회 이상 재방문되어 종료")
                    break

                no = row_step.get("no", f"{index+1}")

                # --- jump_if_visible/visible_if 가드 ---
                jump_if = str(row_step.get("jump_if_visible", row_step.get("visible_if",""))).strip()
                jump_type = str(row_step.get("jump_if_visible_type", row_step.get("visible_if_type","text"))).strip().lower()
                if jump_if:
                    try:
                        if jump_type == "text" and jump_if not in self.driver.page_source:
                            log_skip(f"텍스트 '{jump_if}' 없음 → Step({no}) SKIP")
                            k += 1
                            continue
                        elif jump_type == "xpath" and not self.driver.find_elements(By.XPATH, jump_if):
                            log_skip(f"요소 '{jump_if}' 없음 → Step({no}) SKIP")
                            k += 1
                            continue
                    except Exception as e:
                        log_warn(f"jump_if_visible 평가 오류: {e}")
                        k += 1
                        continue

                # --- 액션 실행 ---
                try:
                    name_raw = row_step.get("name", "")
                    try:
                        name_str = "" if pd.isna(name_raw) else str(name_raw).strip()
                    except Exception:
                        name_str = ""
                    if not name_str or name_str.lower() == "nan":
                        name_str = f"step_{index+1}"

                    success = self.perform_action(
                        action=row_step["action"],
                        by=row_step["by"],
                        value=row_step["value"],
                        name=name_str,
                        sleep_time=row_step.get("sleep", 0),
                        skip_on_error=row_step.get("skip_on_error", "N"),
                        visible_if_type=row_step.get("visible_if_type", "text"),
                        visible_if=row_step.get("visible_if", ""),
                        mandatory=row_step.get("mandatory", "Y"),
                        input_text=row_step.get("input_text", ""),
                        no=no
                    )
                except Exception:
                    self.update_config_cell(idx, "last_failed_step", str(no))
                    raise

                # --- 점프 처리 (정/역방향 모두 허용) ---
                jumped = False
                jump_to = row_step.get("jump_to_no", "")
                if success and pd.notna(jump_to):
                    try:
                        target_no = str(int(float(jump_to))).strip()
                        match = df[df["no_str"] == target_no].index
                        if not match.empty:
                            jump_index = match[0]
                            if jump_index != index:  # 자기자신 점프만 금지
                                direction = "forward" if jump_index > index else "backward"
                                log_info(f"JUMP({direction}) Step {no} → jump_to_no {target_no}")
                                if jump_index in indexes:
                                    k = indexes.index(jump_index)
                                    jumped = True
                                else:
                                    log_warn(f"jump_to_no {target_no}가 현재 실행 범위 밖이므로 무시")
                    except Exception as e:
                        log_fail(f"jump_to_no 처리 오류: {e}")

                if jumped:
                    continue  # 점프 위치에서 다음 사이클
                else:
                    k += 1     # 일반 전진


            self.update_config_cell(idx, "last_failed_step", "")





    def update_config_cell(self, row_idx, column, value):
        try:
            wb = load_workbook(self.action_file)
            ws = wb["Config"]

            # 열 이름 → 열 번호 매핑
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            col_index = None
            for i, col_name in enumerate(header_row, start=1):
                if col_name == column:
                    col_index = i
                    break

            if col_index is None:
                log_warn(f"[SKIP] Config 시트에 '{column}' 열 없음")
                return

            # 셀 업데이트
            ws.cell(row=row_idx + 2, column=col_index).value = value  # +2: 헤더 1행 + 0-based 인덱스
            wb.save(self.action_file)
            wb.close()
            log_ok(f"[UPDATE] Config[{row_idx}]['{column}'] → '{value}' 기록됨 (openpyxl)")

        except Exception as e:
            log_fail(f"Config 시트(openpyxl) 업데이트 실패: {e}")


    # Config 시트의 'test_step_scope' 문자열을 index 리스트로 변환
    def parse_scope_string(self, scope_str, df):
        result_indexes = []
        no_str_series = df["no"].apply(lambda x: str(int(float(x))) if pd.notna(x) else "")
        scope_str = str(scope_str).strip()

        for token in scope_str.split(","):
            token = token.strip()
            if "~" in token:
                try:
                    start, end = token.split("~")
                    start_idx = no_str_series[no_str_series == start.strip()].index
                    end_idx = no_str_series[no_str_series == end.strip()].index
                    if not start_idx.empty and not end_idx.empty:
                        result_indexes.extend(range(start_idx[0], end_idx[0] + 1))
                except Exception:
                    continue
            else:
                match = no_str_series[no_str_series == token].index
                if not match.empty:
                    result_indexes.append(match[0])

        return sorted(set(result_indexes))




    def tearDown(self):
        try:
            self.driver.quit()
        except Exception as e:
            log_warn(f"[tearDown] driver.quit() 실패: {e}")
        finally:
            if hasattr(sys.stdout, 'flush'): sys.stdout.flush()
            sys.stdout = sys.__stdout__
            sys.stderr = sys.__stderr__
